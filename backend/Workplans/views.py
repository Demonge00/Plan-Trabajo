import datetime
import openpyxl
from rest_framework.response import Response
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from django.contrib.auth import get_user_model
from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
from django.core.files.storage import default_storage
from django.db.utils import IntegrityError
from django.core.mail import send_mail
from django.conf import settings
from Workplans.models import Workplans, Activity
import xlrd
import sys
import re
from django.utils.crypto import get_random_string
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework.permissions import IsAuthenticated
from Workplans.serialisers import UserSerializer, WorkplanSerializer
from rest_framework_simplejwt.tokens import RefreshToken

User = get_user_model()
monthdays = ['', 31, [28, 29], 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]


def turnParser(turno):
    turn = turno.strip()
    int_turn = float(turn[0]+turn[1]+'.'+turn[3]+turn[4])
    if (int_turn <= 9.5):
        return 1
    elif (int_turn < 11):
        return 2
    elif (int_turn <= 12.5):
        return 3
    elif (int_turn < 14):
        return 4
    elif (int_turn <= 15.5):
        return 5
    else:
        return 6


class MyTokenObtainPairSerializer(TokenObtainPairSerializer):
    @classmethod
    def get_token(cls, user):
        token = super().get_token(user)
        token['name'] = user.name
        return token


class MyTokenObtainPairView(TokenObtainPairView):
    serializer_class = MyTokenObtainPairSerializer


@api_view(['POST'])
def registerUser(request):
    verify_secret = get_random_string(length=32)
    if 'password' in request.data and 'email' in request.data and 'name' in request.data:
        try:
            user = User.objects.create_user(
                email=request.data['email'],
                name=request.data['name'],
                password=request.data['password'],
                is_active=False,
                is_staff=False,
                verification_secret=verify_secret,
            )
        except IntegrityError:
            return Response({'message': 'User alredy registered'}, status=status.HTTP_400_BAD_REQUEST)
        except:
            return Response({'message': 'User could not be registered'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            # send_mail(
            #     f'Verifica tu cuenta de usuario para {settings.WEB_SITE_NAME}',
            #     f'Para verificar tu cuenta en {
            #         settings.WEB_SITE_NAME}, ve a settings.VERIFICATION_URL}{verify_secret}',
            #     settings.SENDER_EMAIL,
            #     [request.data['email']],
            #     fail_silently=False,
            #     html_message=f'Porfavor ve a <a href="{
            #         settings.VERIFICATION_URL}/{verify_secret}">este email</a> para verificar tu cuenta.',
            # )
            user.sent_verification_email = True
            user.save()
        except:
            print("send mail exceptioin:", sys.exc_info())
            return Response({"message": "Cannot send email"}, status=status.HTTP_400_BAD_REQUEST)
        print(settings.VERIFICATION_URL)
        return Response({'html': f'Porfavor ve a <a href="{
            settings.VERIFICATION_URL}/{verify_secret}">este link</a> para verificar tu cuenta.'
        }, status=status.HTTP_200_OK)
    return Response({'message': 'Missing info'}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
def verifyUser(request, verification_secret):
    try:
        user = User.objects.get(verification_secret=verification_secret)
        user.verificated = True
        user.is_active = True
        user.save()
        return Response({'message': 'user_registered'}, status=status.HTTP_200_OK)
    except:
        return Response({'message': 'Unable to verify email'}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
def handleEmail(request):
    password_secret = get_random_string(length=32)
    try:
        user = User.objects.get(email=request.data['email'])
        user.password_secret = password_secret
        user.save()
        print(f"Porfavor ve a <a href='http://localhost:5173/password/{
              password_secret}'>este link</a> para verificar tu cuenta.")
        return Response(status=status.HTTP_200_OK)
    except:
        return Response({'message': "Mail doesnt exist"}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
def updatePassword(request, password_secret):
    try:
        user = User.objects.get(password_secret=password_secret)
        user.password = make_password(request.data['password'])
        user.save()
        return Response(status=status.HTTP_200_OK)
    except:
        return Response({'message': "Password cant be changed"}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def userInfo(request):
    user = request.user
    serializer = UserSerializer(user, many=False)
    return Response(serializer.data)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def updateInfo(request):
    flag = False
    user = User.objects.get(email=request.user)
    if "name" in request.data:
        user.name = request.data['name']
        flag = True
    if "password" in request.data:
        user.password = make_password(request.data["password"])
        flag = True
    if flag:
        user.save()
    access_token = RefreshToken.for_user(user)
    access_token['name'] = user.name
    return Response({'access': str(access_token.access_token),
                     'refresh': str(access_token)}, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def createWorkplan(request):
    user = request.user
    if (request.FILES):
        name = request.data['name']
        excel = request.FILES['file']
        file_path = default_storage.save('tmp/excel_file.xlsx', excel)
        # Abrir el archivo Excel
        if (re.search(r'\.xls$', name)):
            workbook = xlrd.open_workbook(default_storage.path(file_path))
            sheet = workbook.sheet_by_index(0)
            # Leer los datos de la hoja de cálculo
            data = []
            for row in range(1, sheet.nrows):
                if (re.search(r'profesores', sheet.cell_value(row, 6), re.IGNORECASE)):
                    turno = str(sheet.cell_value(row, 3)).split("-")
                    t1 = turnParser(turno[0])
                    d2 = ''
                    if (str(sheet.cell_value(row, 2))[1] != '-'):
                        d2 = str(sheet.cell_value(row, 2))[1]
                    day = int(str(sheet.cell_value(row, 2)
                                  )[0] + d2)
                    print(day)
                    if (len(turno) > 1):
                        t2 = turnParser(turno[1])
                    else:
                        t2 = turnParser(turno[0])
                    for turnito in range(t1, t2+1):
                        row_data = {
                            'turn': turnito,
                            'day': day,
                            'activity': sheet.cell_value(row, 1)
                        }
                        data.append(row_data)
            print(data)
        elif (re.search(r'\.xlsx$', name)):
            workbook = openpyxl.load_workbook(default_storage.path(file_path))
            worksheet = workbook.active
            data = []
            for row in range(2, worksheet.max_row + 1):
                if re.search(r'profesores', worksheet.cell(row=row, column=7).value, re.IGNORECASE):
                    turno = str(worksheet.cell(
                        row=row, column=4).value).split("-")
                    t1 = turnParser(turno[0])
                    day = int(str(worksheet.cell(row=row, column=3).value)[
                              0] + str(worksheet.cell(row=row, column=3).value)[1])
                    print(day)
                    if len(turno) > 1:
                        t2 = turnParser(turno[1])
                    else:
                        t2 = turnParser(turno[0])
                    for turnito in range(t1, t2 + 1):
                        row_data = {
                            'turn': turnito,
                            'day': day,
                            'activity': worksheet.cell(row=row, column=2).value
                        }
                        data.append(row_data)
            print(data)
        else:
            return Response(status=status.HTTP_400_BAD_REQUEST)

    if (int(request.data["month"]) <= 12 and int(request.data["month"]) > 0):
        date = datetime.date(
            int(request.data["year"]), int(request.data["month"]), 1)
    try:
        work = Workplans.objects.get(date=date)
        work.custom_user.add(user)
        work.save()
        return Response({'message': 'plan conected'}, status=status.HTTP_200_OK)
    except:
        work = Workplans(date=date)
        work.save()
        work.custom_user.add(user)
        work.save()
        if (data):
            for i in data:
                activity = Activity(
                    workplans=work, day=i['day'], turn=i['turn'], activity=i['activity'])
                print(activity)
                activity.save()
        return Response({'message': 'plan created'}, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def listPlans(request):
    user = request.user
    try:
        lista = user.workplans_set.all()
        serializer = WorkplanSerializer(lista, many=True)
        return Response({'list': serializer.data, 'message': 'Done'}, status=status.HTTP_200_OK)
    except:
        return Response({'message': 'Trouble'}, status=status.HTTP_400_BAD_REQUEST)
